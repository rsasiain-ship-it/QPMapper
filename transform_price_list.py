"""
transform_price_list.py
========================
Version: 1.0.0
Spec:    TRANSFORMATION_SPEC.md (keep these two files in sync)

Converts raw vendor/manufacturer price lists into a single normalized
"Quote Data Template" sheet.

Supported source layouts
-------------------------
1. LATTICE   (e.g. Atos Medical)
   - Quick-reference lattice: repeated (Ref#, Price) column-pair blocks.
   - Detailed section lower on the sheet with Description + U/M, keyed on Ref#.
   - Lattice is the authoritative part+price universe; detailed section is
     LEFT-JOINed on to fill Description/UOM where available.

2. TABULAR   (e.g. Tracoe)
   - One row per part, fixed column positions.

Adding a new vendor
--------------------
Add an entry to LAYOUTS (see spec §6) rather than special-casing the main loop.

Usage
-----
    python transform_price_list.py <input_workbook.xlsx> <output_workbook.xlsx>

The script expects a small "layout manifest" mapping each source sheet name
to a manufacturer name + layout type. Edit SHEET_MANIFEST below (or pass
--manifest a JSON file) to describe the sheets in your workbook.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass, field
from typing import Optional

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Output schema (spec §1) — exact column order, do not reorder.
# ---------------------------------------------------------------------------
OUTPUT_COLUMNS = [
    "ManufacturerName",
    "Manufacturer Catalog Number",
    "Manufacturer Catalog Description",
    "Proposed UOM",
    "Proposed UOM Quantity",
    "Proposed UOM Price",
    "Proposed Purchase Quantity",
]

# ---------------------------------------------------------------------------
# Layout configs (spec §2). Add new vendors here, not in the parsing loop.
# ---------------------------------------------------------------------------

LATTICE_BLOCKS = [(1, 2), (5, 6), (9, 10), (12, 14), (16, 17), (22, 24)]

LATTICE_DETAILED_COLS = {
    # 0-indexed offsets FROM COLUMN B (i.e. col_index = B_INDEX + offset)
    "ref": 0,          # col B
    "description": 2,  # col D
    "uom": 16,          # col R
    "price": 19,        # col U
}
_B_INDEX = 1  # column B is index 1 (0-indexed, col A = 0)

TABULAR_COLS = {
    # 0-indexed from col A
    "ref": 1,           # col B
    "sizes": 3,          # col D
    "description": 5,    # col F
    "uom": 8,             # col I
    "pc_per_uom": 10,      # col K
    "price": 11,            # col L
}

LAYOUTS = {
    "lattice": {
        "blocks": LATTICE_BLOCKS,
        "detailed_cols": LATTICE_DETAILED_COLS,
        "detailed_col_offset": _B_INDEX,
    },
    "tabular": {
        "cols": TABULAR_COLS,
    },
}

# ---------------------------------------------------------------------------
# Sheet manifest: which sheet -> which manufacturer + layout.
# Edit this (or supply --manifest some.json) per workbook you run.
# ---------------------------------------------------------------------------
SHEET_MANIFEST: dict[str, dict] = {
    # "Sheet1": {"manufacturer": "Atos Medical", "layout": "lattice"},
    # "Sheet2": {"manufacturer": "Tracoe", "layout": "tabular"},
}


# ---------------------------------------------------------------------------
# Spec §4 — Row validity / junk filter
# ---------------------------------------------------------------------------
NUM_RE = re.compile(r'^\$?\d[\d,]*\.?\d*$')


def is_price(p) -> bool:
    if isinstance(p, (int, float)):
        return True
    if p is None:
        return False
    return bool(NUM_RE.match(str(p).strip()))


def to_price(p) -> float:
    if isinstance(p, (int, float)):
        return float(p)
    return float(str(p).replace('$', '').replace(',', ''))


# ---------------------------------------------------------------------------
# Spec §5 — UOM parsing
# ---------------------------------------------------------------------------
def normalize_unit(u: str) -> str:
    if not u:
        return ""
    l = str(u).lower()
    if l.startswith("pc"):
        return "PC"
    if l.startswith("set"):
        return "SET"
    if l.startswith("kit"):
        return "KIT"
    if l.startswith("box"):
        return "BOX"
    return str(u).upper()


def split_uom(uom_str: str):
    """'1 pc' -> (1, 'PC'); 'BOX' -> ('', 'BOX')"""
    if not uom_str:
        return "", ""
    m = re.match(r'^(\d+)\s*(.*)$', str(uom_str).strip())
    if m:
        qty = int(m.group(1))
        unit = m.group(2).strip() or uom_str
    else:
        qty, unit = "", uom_str
    return qty, normalize_unit(unit)


# ---------------------------------------------------------------------------
# Row model
# ---------------------------------------------------------------------------
@dataclass
class QuoteRow:
    manufacturer: str
    catalog_number: str
    description: str = ""
    uom: str = ""
    uom_qty: object = ""
    uom_price: Optional[float] = None
    purchase_qty: object = ""  # always blank per spec §3

    def as_list(self) -> list:
        return [
            self.manufacturer,
            self.catalog_number,
            self.description,
            self.uom,
            self.uom_qty,
            self.uom_price,
            self.purchase_qty,
        ]


def clean_catalog_number(raw: str) -> str:
    """Strip a leading 'REF ' (case-insensitive); keep the rest verbatim."""
    if raw is None:
        return ""
    s = str(raw).strip()
    return re.sub(r'^REF\s+', '', s, flags=re.IGNORECASE)


# ---------------------------------------------------------------------------
# Lattice layout parser (spec §2a)
# ---------------------------------------------------------------------------
def parse_lattice_sheet(ws: Worksheet, manufacturer: str) -> list[QuoteRow]:
    layout = LAYOUTS["lattice"]

    # Pass 1: walk every lattice block on every row -> authoritative (ref, price)
    lattice_pairs: dict[str, float] = {}
    for row in ws.iter_rows(values_only=True):
        for ref_idx, price_idx in layout["blocks"]:
            if ref_idx >= len(row) or price_idx >= len(row):
                continue
            ref, price = row[ref_idx], row[price_idx]
            if ref is None or not is_price(price):
                continue
            ref_clean = clean_catalog_number(ref)
            if ref_clean:
                lattice_pairs[ref_clean] = to_price(price)

    # Pass 2: walk the detailed section -> Description + U/M keyed on Ref#
    detail_map: dict[str, dict] = {}
    offset = layout["detailed_col_offset"]
    cols = layout["detailed_cols"]
    for row in ws.iter_rows(values_only=True):
        ref_idx = offset + cols["ref"]
        if ref_idx >= len(row):
            continue
        ref = row[ref_idx]
        if ref is None:
            continue
        ref_clean = clean_catalog_number(ref)
        if not ref_clean or ref_clean not in lattice_pairs:
            continue  # detailed section only enriches parts already in lattice
        desc_idx = offset + cols["description"]
        uom_idx = offset + cols["uom"]
        description = row[desc_idx] if desc_idx < len(row) else None
        uom_raw = row[uom_idx] if uom_idx < len(row) else None
        if description or uom_raw:
            detail_map[ref_clean] = {
                "description": description,
                "uom": uom_raw,
            }

    # Build output rows: lattice is authoritative universe, LEFT-JOIN details.
    rows = []
    for ref_clean, price in lattice_pairs.items():
        detail = detail_map.get(ref_clean, {})
        uom_qty, uom_unit = split_uom(detail.get("uom") or "")
        rows.append(
            QuoteRow(
                manufacturer=manufacturer,
                catalog_number=ref_clean,
                description=(detail.get("description") or "").strip()
                if detail.get("description")
                else "",
                uom=uom_unit,
                uom_qty=uom_qty,
                uom_price=price,
            )
        )
    return rows


# ---------------------------------------------------------------------------
# Tabular layout parser (spec §2b)
# ---------------------------------------------------------------------------
def parse_tabular_sheet(ws: Worksheet, manufacturer: str) -> list[QuoteRow]:
    cols = LAYOUTS["tabular"]["cols"]
    rows = []
    for row in ws.iter_rows(values_only=True):
        max_idx = max(cols.values())
        if max_idx >= len(row):
            continue
        ref = row[cols["ref"]]
        price = row[cols["price"]]
        if ref is None or not is_price(price):
            continue

        ref_clean = clean_catalog_number(ref)
        if not ref_clean:
            continue

        description = row[cols["description"]] or ""
        sizes = row[cols["sizes"]]
        if sizes:
            description = f"{description} ({sizes})".strip()

        uom_raw = row[cols["uom"]]
        uom_unit = normalize_unit(uom_raw) if uom_raw else ""
        pc_per_uom = row[cols["pc_per_uom"]]

        rows.append(
            QuoteRow(
                manufacturer=manufacturer,
                catalog_number=ref_clean,
                description=str(description).strip(),
                uom=uom_unit,
                uom_qty=pc_per_uom if pc_per_uom is not None else "",
                uom_price=to_price(price),
            )
        )
    return rows


PARSERS = {
    "lattice": parse_lattice_sheet,
    "tabular": parse_tabular_sheet,
}


# ---------------------------------------------------------------------------
# Orchestration
# ---------------------------------------------------------------------------
def transform_workbook(input_path: str, output_path: str, manifest: dict) -> int:
    wb = load_workbook(input_path, data_only=True)
    all_rows: list[QuoteRow] = []

    for sheet_name, config in manifest.items():
        if sheet_name not in wb.sheetnames:
            print(f"[warn] sheet '{sheet_name}' not found in {input_path}, skipping")
            continue
        ws = wb[sheet_name]
        layout = config.get("layout")
        manufacturer = config.get("manufacturer")
        parser = PARSERS.get(layout)
        if parser is None:
            print(f"[warn] unknown layout '{layout}' for sheet '{sheet_name}', skipping")
            continue
        rows = parser(ws, manufacturer)
        print(f"[info] {sheet_name} ({layout}): {len(rows)} rows")
        all_rows.extend(rows)

    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Quote Data Template"
    out_ws.append(OUTPUT_COLUMNS)
    for r in all_rows:
        out_ws.append(r.as_list())

    out_wb.save(output_path)
    print(f"[info] wrote {len(all_rows)} total rows -> {output_path}")
    return len(all_rows)


def main():
    parser = argparse.ArgumentParser(description="Transform vendor price lists into the Quote Data Template.")
    parser.add_argument("input", help="Path to source workbook (.xlsx)")
    parser.add_argument("output", help="Path to write the Quote Data Template (.xlsx)")
    parser.add_argument(
        "--manifest",
        help="Path to a JSON file mapping sheet name -> {manufacturer, layout}. "
             "Overrides SHEET_MANIFEST in this file if provided.",
    )
    args = parser.parse_args()

    manifest = SHEET_MANIFEST
    if args.manifest:
        with open(args.manifest) as f:
            manifest = json.load(f)

    if not manifest:
        print(
            "[error] no sheet manifest configured. Edit SHEET_MANIFEST in this "
            "script or pass --manifest path/to/manifest.json",
            file=sys.stderr,
        )
        sys.exit(1)

    transform_workbook(args.input, args.output, manifest)


if __name__ == "__main__":
    main()
